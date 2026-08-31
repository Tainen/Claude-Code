#!/usr/bin/env python3
"""商談管理シートを「1企業1行（サービス横並び）」から「1サービス1行（縦持ち）」へ変換する。

旧フォーマット（列 A〜CP / 1企業 = 1行）
  A〜R              企業・商談の共通情報
  S〜Z              1つ目に提案しているサービス
  AA〜AH            2つ目のサービス      （AI〜AL: そのサービスのネクスト情報）
  AM〜AT            3つ目のサービス      （AU〜AX: 同上）
  AY〜BF            4つ目のサービス      （BG〜BJ: 同上）
  BK〜BR            5つ目のサービス      （BS〜BV: 同上）
  BW〜CD            6つ目のサービス      （CM〜CP: 同上）
  CE〜CL            7つ目のサービス

新フォーマット（列 A〜AE / 1サービス = 1行）
  A〜R              企業・商談の共通情報（同一企業の全行で同じ値を複製）
  S〜Z              そのサービスの詳細（商談者／商談日／商談回数／サービス詳細／
                    フェーズ／提案金額／申込期日／失注理由）
  AA〜AE            補助列（サービスNo と、旧シートでサービス2以降に紐づいていた
                    ネクスト期日／ネクストアクション／担当者メモ／文字起こし）

使い方:
    python3 restructure_shodan.py 入力.xlsx 出力.xlsx [シート名]
"""

import datetime
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string as ci
from openpyxl.utils import get_column_letter as gl

SHEET_NAME = "商談管理"

# 共通列（A〜R）
COMMON_LAST = ci("R")

# サービスブロックの先頭列。1ブロック = 8列。
SERVICE_BLOCKS = ["S", "AA", "AM", "AY", "BK", "BW", "CE"]
SERVICE_WIDTH = 8

# 各サービスに紐づくネクスト情報（4列）。サービス1の分は共通列 O〜R と同じ内容
# なので重複を避けて None にしてある。サービス7には対応するブロックが無い。
SERVICE_EXTRA = [None, "AI", "AU", "BG", "BS", "CM", None]
EXTRA_WIDTH = 4

SERVICE_HEADERS = [
    "商談者",
    "商談日",
    "商談回数",
    "サービス詳細",
    "フェーズ",
    "提案金額",
    "申込期日",
    "失注理由",
]
EXTRA_HEADERS = [
    "サービスNo",
    "ネクスト期日\n(サービス別)",
    "ネクストアクション\n(サービス別)",
    "担当者メモ\n(サービス別)",
    "文字起こし\n(サービス別)",
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SERVICE_FILL = PatternFill("solid", fgColor="FCE4D6")
EXTRA_FILL = PatternFill("solid", fgColor="EDEDED")


DATE_FORMAT = 'yyyy"年"m"月"d"日"'


def write(out, row, col, value, number_format):
    """値を書き込み、日付なら書式が崩れないよう日付書式を優先する。"""
    cell = out.cell(row, col, value)
    if isinstance(value, (datetime.datetime, datetime.date)):
        cell.number_format = DATE_FORMAT
    else:
        cell.number_format = number_format
    return cell


def block_is_empty(ws, row, start_col, width):
    return all(
        ws.cell(row, start_col + i).value in (None, "") for i in range(width)
    )


def read_source(path):
    """値のみのブックと書式用のブックを返す。"""
    values = openpyxl.load_workbook(path, data_only=True)
    formats = openpyxl.load_workbook(path)
    return values, formats


def convert(src_path, dst_path, sheet_name=SHEET_NAME):
    vwb, fwb = read_source(src_path)
    src = vwb[sheet_name]
    fmt = fwb[sheet_name]

    last_row = max(
        (r for r in range(2, src.max_row + 1) if src.cell(r, 3).value not in (None, "")),
        default=1,
    )

    out_wb = openpyxl.Workbook()
    out = out_wb.active
    out.title = sheet_name

    # ---- ヘッダー ----
    for c in range(1, COMMON_LAST + 1):
        cell = out.cell(1, c, src.cell(1, c).value)
        cell.fill = HEADER_FILL
    for i, name in enumerate(SERVICE_HEADERS):
        cell = out.cell(1, ci("S") + i, name)
        cell.fill = SERVICE_FILL
    for i, name in enumerate(EXTRA_HEADERS):
        cell = out.cell(1, ci("AA") + i, name)
        cell.fill = EXTRA_FILL
    for c in range(1, ci("AE") + 1):
        cell = out.cell(1, c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # 旧シートの表示形式を引き継ぐための番号書式
    common_fmt = [fmt.cell(2, c).number_format for c in range(1, COMMON_LAST + 1)]
    service_fmt = [
        fmt.cell(2, ci("S") + i).number_format for i in range(SERVICE_WIDTH)
    ]
    extra_fmt = ["General"] + [
        fmt.cell(2, ci("AI") + i).number_format for i in range(EXTRA_WIDTH)
    ]
    # 旧シートでは長文列に誤って数値書式が付いているため、文字列列は General に戻す
    text_cols = {ci("N"), ci("P"), ci("Q"), ci("R")}
    for c in text_cols:
        common_fmt[c - 1] = "General"
    service_fmt[3] = service_fmt[4] = "General"
    extra_fmt[2] = extra_fmt[3] = extra_fmt[4] = "General"

    out_row = 1
    companies = 0
    for r in range(2, last_row + 1):
        if src.cell(r, 3).value in (None, ""):
            continue
        companies += 1

        filled = [
            idx
            for idx, start in enumerate(SERVICE_BLOCKS)
            if not block_is_empty(src, r, ci(start), SERVICE_WIDTH)
        ]
        # サービスが1つも入力されていない企業も1行は残す
        targets = filled if filled else [None]

        for idx in targets:
            out_row += 1
            for c in range(1, COMMON_LAST + 1):
                write(out, out_row, c, src.cell(r, c).value, common_fmt[c - 1])

            if idx is None:
                continue

            start = ci(SERVICE_BLOCKS[idx])
            for i in range(SERVICE_WIDTH):
                write(
                    out, out_row, ci("S") + i, src.cell(r, start + i).value,
                    service_fmt[i],
                )

            out.cell(out_row, ci("AA"), idx + 1)
            extra = SERVICE_EXTRA[idx]
            if extra:
                estart = ci(extra)
                for i in range(EXTRA_WIDTH):
                    write(
                        out, out_row, ci("AB") + i, src.cell(r, estart + i).value,
                        extra_fmt[i + 1],
                    )

    # ---- 体裁 ----
    for col, width in (
        ("A", 6), ("B", 9), ("C", 28), ("D", 14), ("E", 12), ("F", 12),
        ("G", 8), ("H", 8), ("I", 13), ("J", 24), ("K", 14), ("L", 13),
        ("M", 13), ("N", 40), ("O", 11), ("P", 18), ("Q", 24), ("R", 40),
        ("S", 9), ("T", 12), ("U", 9), ("V", 28), ("W", 16), ("X", 10),
        ("Y", 12), ("Z", 20), ("AA", 9), ("AB", 12), ("AC", 18), ("AD", 24),
        ("AE", 40),
    ):
        out.column_dimensions[col].width = width
    out.freeze_panes = "D2"
    out.auto_filter.ref = f"A1:{gl(ci('AE'))}{out_row}"

    out_wb.save(dst_path)
    return companies, out_row - 1


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        return 1
    sheet = sys.argv[3] if len(sys.argv) > 3 else SHEET_NAME
    companies, rows = convert(sys.argv[1], sys.argv[2], sheet)
    print(f"{companies} 企業 -> {rows} 行を書き出しました: {sys.argv[2]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
